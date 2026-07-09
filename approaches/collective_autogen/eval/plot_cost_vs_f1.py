"""Plot partial_F1 vs (a) USD cost, (b) total tokens, (c) wall-clock seconds —
one labelled point per system.

Inputs:
- F1: read from `eval/results/eval_summary.xlsx`, sheet `GT3_Maarten_aggregate`,
  column `mean_partial_f1`.
- Tokens / time: each system's latest `summary.json` under `eval/results/`,
  fields `usage_total.{prompt_tokens, completion_tokens}` and
  `elapsed_total_s` / `elapsed_s` (hybrid suites use `elapsed_s` for the whole
  worker invocation).
- Prices: `eval/model_prices.json` — USD per 1M input/output tokens.

Map of system → suite directory glob is configurable on the CLI; defaults
correspond to the systems on the aggregate sheet.

Output: three PNGs under `eval/results/cost_vs_f1/`.
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import matplotlib.pyplot as plt
from openpyxl import load_workbook

REPO = Path(__file__).resolve().parent.parent
RESULTS = REPO / "eval" / "results"
XLSX = RESULTS / "eval_summary.xlsx"
PRICES = REPO / "eval" / "model_prices.json"
OUT_DIR = RESULTS / "cost_vs_f1"

# Sibling testje project (ReactionMiner) — single-agent runs live there as
# benchmark_runs/<run>/<stem>.meta.json sidecars (one per image).
# REPO is `.../thesis/5.Random_tests/Collective_autogen`; testje sits next to
# `thesis` under `Documents/`.
TESTJE_ROOT = REPO.parent.parent.parent / "testje"
TESTJE_RUNS = TESTJE_ROOT / "benchmark_runs"


# Each system on the aggregate sheet is matched by one of two rule shapes:
#   ("thesis", dir_glob, model_match)
#       — find latest suite_*/summary.json with usage_total under eval/results/
#   ("testje", run_dir_name, fallback_model_for_pricing)
#       — sum <stem>.meta.json sidecars under TESTJE_RUNS/<run_dir_name>/
#       — falls back to _costs_summary.json if sidecars are absent; in that
#         case fallback_model_for_pricing is used to look up the price.
DEFAULT_SUITE_RULES: dict = {
    "multi_agent":                                ("thesis", "suite_*",            None),
    "ChemEagle":                                  ("thesis", "chemeagle_2*",       None),
    "ChemEagle_Hybrid":                           ("thesis", "chemeagle_hybrid_*", None),
    "single_sdk_agent (Opus 4.7)":                ("testje", "run01_opus4.7",          "claude-opus-4-7"),
    "single_sdk_agent (GPT-5.4)":                 ("testje", "run_gpt54",              "gpt-5.4"),
    "single_sdk_agent (Gemini 3 Flash Preview)":  ("testje", "run_gemini3flash",       "gemini-3-flash-preview"),
    "single_sdk_agent (Gemini 3.1 Pro Preview)":  ("testje", "run_gemini31propreview", "gemini-3-pro-preview"),
}


def _suite_model(suite_dir: Path) -> str | None:
    sj = suite_dir / "summary.json"
    if not sj.exists():
        return None
    try:
        d = json.loads(sj.read_text())
    except Exception:
        return None
    return d.get("vision_model") or d.get("deployment") or (
        (d.get("usage_total") or {}).get("model")
    )


def _latest_suite_dir(dir_glob: str, model_match: list[str] | None) -> Path | None:
    matches = [p for p in RESULTS.glob(dir_glob) if p.is_dir()]
    if model_match:
        wanted = [s.lower() for s in model_match]
        matches = [
            p for p in matches
            if (m := _suite_model(p)) and all(w in m.lower() for w in wanted)
        ]
    if not matches:
        return None
    return max(matches, key=lambda p: p.stat().st_mtime)


def _load_partial_f1() -> dict[str, float]:
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["GT3_Maarten_aggregate"]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    sys_idx = header.index("system")
    f1_idx = header.index("mean_partial_f1")
    return {row[sys_idx]: row[f1_idx] for row in rows[1:] if row and row[sys_idx]}


def _load_prices() -> dict[str, dict[str, float]]:
    if not PRICES.exists():
        return {}
    return json.loads(PRICES.read_text()).get("models", {})


def _price_for(model: str | None, prices: dict) -> tuple[float, float] | None:
    """Return (input_per_m, output_per_m) USD prices for `model`, or None."""
    if not model:
        return None
    if model in prices:
        e = prices[model]
        return float(e["input_per_m"]), float(e["output_per_m"])
    lo = model.lower()
    for k, e in prices.items():
        if k.lower() in lo or lo in k.lower():
            return float(e["input_per_m"]), float(e["output_per_m"])
    return None


def _load_suite_metrics(suite_dir: Path) -> dict | None:
    sj = suite_dir / "summary.json"
    if not sj.exists():
        return None
    try:
        d = json.loads(sj.read_text())
    except Exception:
        return None
    usage = d.get("usage_total") or {}
    elapsed = d.get("elapsed_total_s")
    if elapsed is None:
        elapsed = d.get("elapsed_s")
    if elapsed is None:
        elapsed = sum(float(p.get("elapsed_s") or 0) for p in d.get("per_image", []))
    per_image = d.get("per_image") or []
    n_total = len(per_image)
    n_ok = sum(
        1 for p in per_image
        if (p.get("rc") == 0 or p.get("ok") is True) and not p.get("error")
    )
    return {
        "model": usage.get("model") or d.get("vision_model") or d.get("deployment"),
        "prompt_tokens": int(usage.get("prompt_tokens") or 0),
        "completion_tokens": int(usage.get("completion_tokens") or 0),
        "total_tokens": int(usage.get("total_tokens") or 0),
        "elapsed_s": float(elapsed or 0),
        "suite_dir": str(suite_dir.relative_to(REPO)),
        "n_ok": n_ok,
        "n_total": n_total,
    }


def _load_testje_run(run_dir_name: str, fallback_model: str | None = None) -> dict | None:
    """Sum every `<stem>.meta.json` sidecar in a testje benchmark_runs subdir.

    Falls back to `benchmark_runs/_costs_summary.json` (produced by
    `scripts/summarize_costs.py`) for runs whose sidecars were never written
    — Gemini Flash / Pro Preview were captured from /tmp log lines only.

    Returns the same shape as `_load_suite_metrics`, or None if neither has it."""
    run_dir = TESTJE_RUNS / run_dir_name
    if run_dir.is_dir():
        sidecars = sorted(run_dir.glob("*.meta.json"))
        if sidecars:
            p_total = c_total = 0
            elapsed_total = 0.0
            model = None
            for s in sidecars:
                try:
                    d = json.loads(s.read_text())
                except Exception:
                    continue
                p_total += int(d.get("input_tokens") or 0)
                c_total += int(d.get("output_tokens") or 0)
                elapsed_total += float(d.get("elapsed_s") or 0)
                if model is None and d.get("model"):
                    model = d["model"]
            if p_total or c_total:
                return {
                    "model": model or fallback_model,
                    "prompt_tokens": p_total,
                    "completion_tokens": c_total,
                    "total_tokens": p_total + c_total,
                    "elapsed_s": elapsed_total,
                    "suite_dir": f"<testje>/benchmark_runs/{run_dir_name}",
                    "n_ok": len(sidecars),
                    "n_total": 16,
                }

    # Fallback: aggregated _costs_summary.json (one entry per run dir)
    costs_path = TESTJE_RUNS / "_costs_summary.json"
    if not costs_path.exists():
        return None
    try:
        costs = json.loads(costs_path.read_text())
    except Exception:
        return None
    e = costs.get(run_dir_name)
    if not e or not (e.get("input_tokens_total") or e.get("output_tokens_total")):
        return None
    p = int(e.get("input_tokens_total") or 0)
    c = int(e.get("output_tokens_total") or 0)
    return {
        "model": fallback_model,
        "prompt_tokens": p,
        "completion_tokens": c,
        "total_tokens": p + c,
        "elapsed_s": float(e.get("elapsed_s_total") or 0),
        "suite_dir": f"<testje>/benchmark_runs/_costs_summary.json[{run_dir_name}]",
        "n_ok": int(e.get("n") or 0),
        "n_total": 16,
    }


def _scatter(points: list[dict], x_key: str, x_label: str, title: str,
             out_path: Path, log_x: bool) -> None:
    fig, ax = plt.subplots(figsize=(8, 6))
    for p in points:
        ax.scatter(p[x_key], p["partial_f1"], s=80, zorder=3)
        ax.annotate(p["label"], (p[x_key], p["partial_f1"]),
                    xytext=(8, 4), textcoords="offset points", fontsize=9)
    if log_x:
        ax.set_xscale("log")
    ax.set_xlabel(x_label)
    ax.set_ylabel("partial F1 (mean over benchmark images)")
    ax.set_title(title)
    ax.grid(True, alpha=0.3, zorder=0)
    ax.set_ylim(0, 1)
    fig.tight_layout()
    fig.savefig(out_path, dpi=150)
    plt.close(fig)
    print(f"  wrote {out_path.relative_to(REPO)}")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--out-dir", type=Path, default=OUT_DIR)
    parser.add_argument(
        "--suite", action="append", default=[], metavar="SYSTEM=GLOB[:MODEL_HINT]",
        help="Override default suite rule for a system label, e.g. "
             "'single_sdk_agent (Opus 4.7)=single_20260601_*'. The optional "
             "':MODEL_HINT' filters by substring match on vision_model/"
             "deployment. May repeat.",
    )
    args = parser.parse_args()

    rules = dict(DEFAULT_SUITE_RULES)
    for spec in args.suite:
        if "=" not in spec:
            print(f"--suite must be SYSTEM=GLOB[:MODEL_HINT] (got {spec!r})",
                  file=sys.stderr)
            return 2
        sysname, rest = spec.split("=", 1)
        if ":" in rest:
            glob, hint = rest.split(":", 1)
            rules[sysname.strip()] = ("thesis", glob.strip(), [hint.strip()])
        else:
            rules[sysname.strip()] = ("thesis", rest.strip(), None)

    f1_by_system = _load_partial_f1()
    prices = _load_prices()

    points: list[dict] = []
    skipped: list[str] = []
    for system, rule in rules.items():
        f1 = f1_by_system.get(system)
        if f1 is None:
            skipped.append(f"{system}: not in aggregate sheet")
            continue

        m: dict | None = None
        source_label: str = ""
        if rule[0] == "thesis":
            _, dir_glob, model_match = rule
            d = _latest_suite_dir(dir_glob, model_match)
            if d is None:
                hint = f" with model~{model_match}" if model_match else ""
                skipped.append(f"{system}: no suite matching {dir_glob!r}{hint}")
                continue
            m = _load_suite_metrics(d)
            source_label = d.name
            if m is None or m["total_tokens"] == 0:
                skipped.append(f"{system}: no usage_total in {d.name}")
                continue
        elif rule[0] == "testje":
            _, run_dir_name, fallback_model = rule
            m = _load_testje_run(run_dir_name, fallback_model=fallback_model)
            source_label = f"testje/{run_dir_name}"
            if m is None:
                skipped.append(f"{system}: no usage data for "
                               f"testje/benchmark_runs/{run_dir_name}/ "
                               f"(neither sidecars nor _costs_summary.json)")
                continue
        else:
            skipped.append(f"{system}: unknown rule {rule!r}")
            continue

        prc = _price_for(m["model"], prices)
        if prc is None:
            skipped.append(f"{system}: no price for model {m['model']!r}")
            continue
        in_per_m, out_per_m = prc
        cost_usd = (
            m["prompt_tokens"] * in_per_m + m["completion_tokens"] * out_per_m
        ) / 1_000_000
        n_ok = m.get("n_ok") or 0
        n_total = m.get("n_total") or 16
        # Annotate the system label when not all images succeeded — flags
        # ChemEagle's 12/16 (4 deterministic MolNexTR crashes) and the
        # original Gemini 3.1 Pro 15/16 in the plotted points themselves.
        plot_label = system if n_ok == n_total else f"{system}  ({n_ok}/{n_total})"
        points.append({
            "label": plot_label,
            "system": system,
            "n_ok": n_ok,
            "n_total": n_total,
            "partial_f1": float(f1),
            "cost_usd": cost_usd,
            "total_tokens": m["total_tokens"],
            "elapsed_s": m["elapsed_s"],
            "model": m["model"],
            "suite_dir": m["suite_dir"],
        })
        coverage = "" if n_ok == n_total else f"  [{n_ok}/{n_total}]"
        print(f"  {system}: F1={f1:.3f}  ${cost_usd:.2f}  "
              f"{m['total_tokens']:>10,}tok  {m['elapsed_s']:>6.0f}s{coverage}  "
              f"({source_label}, model={m['model']!r})")

    if skipped:
        print("\nSkipped:")
        for s in skipped:
            print(f"  - {s}")

    if not points:
        print("\nNo points to plot.")
        return 1

    args.out_dir.mkdir(parents=True, exist_ok=True)
    _scatter(points, "cost_usd", "USD cost (16-image benchmark)",
             "Partial F1 vs USD cost", args.out_dir / "f1_vs_cost.png", log_x=True)
    _scatter(points, "total_tokens", "Total tokens (16-image benchmark)",
             "Partial F1 vs token count", args.out_dir / "f1_vs_tokens.png", log_x=True)
    _scatter(points, "elapsed_s", "Wall-clock seconds (16-image benchmark)",
             "Partial F1 vs wall-clock time", args.out_dir / "f1_vs_time.png", log_x=False)

    (args.out_dir / "points.json").write_text(json.dumps(points, indent=2))
    print(f"\n  wrote {(args.out_dir / 'points.json').relative_to(REPO)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
