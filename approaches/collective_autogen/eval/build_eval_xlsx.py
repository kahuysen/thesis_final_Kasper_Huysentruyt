"""Build a single-file Excel workbook of evaluation results.

Sheets:
  1. GT3_Maarten_per_image — per-image scores for multi-agent, ChemEagle,
     ChemEagle_Hybrid on the 16-image GT3_Maarten benchmark.
  2. GT3_Maarten_aggregate — one row per system with mean of each metric.
  3. Multi_agent_phase_trajectory — one row per `suite_*` with mean metrics
     against whichever gold dir the suite was run on.

Usage:
    python eval/build_eval_xlsx.py [--out path/to/file.xlsx]
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from statistics import mean

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO))

from eval.metrics import evaluate  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font, PatternFill  # noqa: E402

RESULTS = REPO / "eval" / "results"
GOLD_GT3 = REPO / "eval" / "Benchmark_kasper_GT3_Maarten" / "ground_truth"
GOLD_OLD = REPO / "eval" / "ground_truth"

GT3_SUITES = {
    "multi_agent": "suite_20260429_123624",
    "ChemEagle":   "chemeagle_20260429_194529",
    "ChemEagle_Hybrid": "chemeagle_hybrid_20260501_143554",
}

METRIC_COLS = [
    ("schema_pass",         lambda m: bool(m["schema_conformance"]["valid"])),
    ("n_reactions_pred",    lambda m: m["reaction_count"]["count"]),
    ("n_reactions_gold",    lambda m: m.get("reaction_count_gold", {}).get("count")),
    ("smiles_valid_rate",   lambda m: m["smiles_validity_rate"]["rate"]),
    ("role_enum_rate",      lambda m: m["role_enum_compliance"]["rate"]),
    ("reactant_iou",        lambda m: m["reactant_iou"]["iou"]),
    ("product_iou",         lambda m: m["product_iou"]["iou"]),
    ("cond_precision",      lambda m: m["condition_coverage"]["precision"]),
    ("cond_recall",         lambda m: m["condition_coverage"]["recall"]),
    ("soft_f1",             lambda m: m["soft_match"]["f1"]),
    ("hard_f1",             lambda m: m["hard_match"]["f1"]),
    ("constitution_f1",     lambda m: m["constitution_match"]["f1"]),
    ("partial_f1",          lambda m: m["partial_match"]["f1"]),
    ("ged",                 lambda m: m["graph_edit_distance"]["avg_ged"]),
]


def _safe(fn, m):
    try:
        return fn(m)
    except Exception:
        return None


def _load_pred(suite_dir: Path, stem: str) -> dict | None:
    """Return prediction record for `stem` from a suite, or None if missing."""
    direct = suite_dir / f"{stem}.converted.json"
    if direct.exists():
        try:
            return json.loads(direct.read_text())
        except Exception:
            return None

    summary_path = suite_dir / "summary.json"
    if not summary_path.exists():
        return None
    summary = json.loads(summary_path.read_text())
    for entry in summary.get("per_image", []):
        if entry.get("stem") != stem:
            continue
        rd = entry.get("run_dir")
        if not rd:
            return None
        rj = REPO / rd / "result.json"
        if not rj.exists() or rj.stat().st_size == 0:
            return None
        try:
            return json.loads(rj.read_text())
        except Exception:
            return None
    return None


def _score_suite_against_gold(suite_dir: Path, gold_dir: Path) -> list[tuple[str, dict | None]]:
    """For every gold file, score the matching prediction. Returns
    [(stem, metrics_or_None), ...]."""
    rows = []
    for gold_path in sorted(gold_dir.glob("*.json")):
        stem = gold_path.stem
        pred = _load_pred(suite_dir, stem)
        if pred is None:
            rows.append((stem, None))
            continue
        try:
            m = evaluate(pred, json.loads(gold_path.read_text()))
        except Exception as e:
            rows.append((stem, {"_error": str(e)}))
            continue
        rows.append((stem, m))
    return rows


def _detect_gold_dir(summary: dict) -> Path | None:
    g = summary.get("gold_dir")
    if g:
        p = REPO / g
        if p.exists():
            return p
    return None


# --- sheet builders --------------------------------------------------------

HEADER_FONT = Font(bold=True)
SECTION_FILL = PatternFill("solid", fgColor="DDDDDD")


def sheet_per_image(wb: Workbook) -> None:
    ws = wb.create_sheet("GT3_Maarten_per_image")
    headers = ["image", "system"] + [c for c, _ in METRIC_COLS]
    ws.append(headers)
    for c in ws[1]:
        c.font = HEADER_FONT

    gold_stems = sorted(p.stem for p in GOLD_GT3.glob("*.json"))
    for stem in gold_stems:
        for system, suite_name in GT3_SUITES.items():
            suite = RESULTS / suite_name
            pred = _load_pred(suite, stem)
            if pred is None:
                ws.append([stem, system] + [None] * len(METRIC_COLS))
                continue
            try:
                m = evaluate(pred, json.loads((GOLD_GT3 / f"{stem}.json").read_text()))
            except Exception as e:
                ws.append([stem, system, f"ERROR: {e}"] + [None] * (len(METRIC_COLS) - 1))
                continue
            ws.append([stem, system] + [_safe(fn, m) for _, fn in METRIC_COLS])

    ws.freeze_panes = "C2"
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 18


def sheet_aggregate(wb: Workbook) -> None:
    ws = wb.create_sheet("GT3_Maarten_aggregate")
    headers = ["system", "n_images_scored", "n_schema_pass"] + [
        f"mean_{c}" for c, _ in METRIC_COLS if c not in ("schema_pass", "n_reactions_pred", "n_reactions_gold")
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = HEADER_FONT

    for system, suite_name in GT3_SUITES.items():
        suite = RESULTS / suite_name
        rows = _score_suite_against_gold(suite, GOLD_GT3)
        scored = [m for _, m in rows if isinstance(m, dict) and "_error" not in m]
        n = len(scored)
        n_pass = sum(1 for m in scored if m["schema_conformance"]["valid"])
        means = []
        for col, fn in METRIC_COLS:
            if col in ("schema_pass", "n_reactions_pred", "n_reactions_gold"):
                continue
            vals = [v for v in (_safe(fn, m) for m in scored) if isinstance(v, (int, float))]
            means.append(round(mean(vals), 4) if vals else None)
        ws.append([system, n, n_pass] + means)

    ws.column_dimensions["A"].width = 22


def sheet_phase_trajectory(wb: Workbook) -> None:
    ws = wb.create_sheet("Multi_agent_phase_trajectory")
    headers = [
        "suite_id", "kind", "vision_model", "mini_model", "use_molnextr",
        "gold_dir", "n_gold", "n_predictions", "n_schema_pass",
    ] + [f"mean_{c}" for c, _ in METRIC_COLS if c not in ("schema_pass", "n_reactions_pred", "n_reactions_gold")]
    ws.append(headers)
    for c in ws[1]:
        c.font = HEADER_FONT

    suite_dirs = sorted([d for d in RESULTS.iterdir() if d.is_dir()])
    for suite in suite_dirs:
        summary_path = suite / "summary.json"
        if not summary_path.exists():
            continue
        try:
            summary = json.loads(summary_path.read_text())
        except Exception:
            continue
        gold_dir = _detect_gold_dir(summary) or GOLD_OLD
        rows = _score_suite_against_gold(suite, gold_dir)
        scored = [m for _, m in rows if isinstance(m, dict) and "_error" not in m]
        n_gold = len(rows)
        n_pred = len(scored)
        n_pass = sum(1 for m in scored if m["schema_conformance"]["valid"])
        means = []
        for col, fn in METRIC_COLS:
            if col in ("schema_pass", "n_reactions_pred", "n_reactions_gold"):
                continue
            vals = [v for v in (_safe(fn, m) for m in scored) if isinstance(v, (int, float))]
            means.append(round(mean(vals), 4) if vals else None)

        ws.append([
            summary.get("suite_id", suite.name),
            summary.get("kind", "multi_agent"),
            summary.get("vision_model"),
            summary.get("mini_model"),
            summary.get("use_molnextr"),
            str(gold_dir.relative_to(REPO)),
            n_gold, n_pred, n_pass,
        ] + means)

    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["F"].width = 48


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default="eval/results/eval_summary.xlsx")
    args = ap.parse_args()

    wb = Workbook()
    wb.remove(wb.active)
    sheet_per_image(wb)
    sheet_aggregate(wb)
    sheet_phase_trajectory(wb)

    out = REPO / args.out
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    main()