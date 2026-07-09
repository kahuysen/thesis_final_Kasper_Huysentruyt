"""Re-score every system's condition_precision/condition_recall in
eval_summary.xlsx using the lenient matcher, and add a `cond_metric_mode`
column to both sheets so the change is documented in-band.

For our two systems we already have lenient values in the file (from the
last run of append_to_eval_xlsx.py). For multi_agent / ChemEagle /
ChemEagle_Hybrid we read each prediction from the suite's run_dir mapping
in summary.json and re-score from raw predictions.

After running, both sheets have:
  - cond_precision, cond_recall columns: lenient values everywhere
  - cond_metric_mode column: "lenient" everywhere
  - aggregate row's means recomputed for the 3 retroactively-rescored systems
"""
from __future__ import annotations

import json
import sys
from pathlib import Path
from statistics import mean

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

COLL = Path("/Users/kasperhuysentruyt/Documents/thesis/5.Random_tests/Collective_autogen")
sys.path.insert(0, str(COLL))
sys.path.insert(0, str(ROOT / "scripts"))

from openpyxl import load_workbook
from eval_via_collective import lenient_condition_coverage

XLSX_PATH    = COLL / "eval" / "results" / "eval_summary.xlsx"
GOLD_DIR     = COLL / "eval" / "ground_truth"
RESULTS_DIR  = COLL / "eval" / "results"

# system → suite-dir name (relative to eval/results/). Mirrors GT3_SUITES
# in build_eval_xlsx.py — these are the systems already in the sheet that
# were originally scored under the strict matcher.
SUITE_FOR_SYSTEM = {
    "multi_agent":      "suite_20260429_123624",
    "ChemEagle":        "chemeagle_20260429_194529",
    "ChemEagle_Hybrid": "chemeagle_hybrid_20260501_143554",
}

# Any system whose name matches this predicate was scored with the lenient
# matcher at append time (see append_to_eval_xlsx.py). Just tag the row;
# don't recompute.
def _is_already_lenient(system: str) -> bool:
    return system.startswith("single_sdk_agent")


def _load_pred(suite_dir: Path, stem: str):
    """Load a per-image prediction from a suite dir.

    Two suite layouts in the wild:
      - multi_agent: summary.json[per_image].run_dir  →  <COLL>/<rd>/result.json
      - ChemEagle*:  summary.json[per_image].native_json (a path) — but the
        post-converted full record lives next to it as `<stem>.converted.json`
        in the suite dir itself.
    Try both.
    """
    summary_path = suite_dir / "summary.json"
    if not summary_path.exists():
        return None
    summary = json.loads(summary_path.read_text())
    for entry in summary.get("per_image", []):
        if entry.get("stem") != stem:
            continue

        # Path A: run_dir/result.json (multi_agent)
        rd = entry.get("run_dir")
        if rd:
            rj = COLL / rd / "result.json"
            if rj.exists() and rj.stat().st_size > 0:
                try:
                    return json.loads(rj.read_text())
                except Exception:
                    pass

        # Path B: <suite>/<stem>.converted.json (ChemEagle, ChemEagle_Hybrid)
        conv = suite_dir / f"{stem}.converted.json"
        if conv.exists() and conv.stat().st_size > 0:
            try:
                return json.loads(conv.read_text())
            except Exception:
                pass

        return None
    return None


def _lenient_for(system: str, stem: str) -> dict | None:
    """Score one (system, image) pair with the lenient matcher."""
    suite = SUITE_FOR_SYSTEM.get(system)
    if not suite:
        return None
    pred = _load_pred(RESULTS_DIR / suite, stem)
    if pred is None:
        return None
    gold_path = GOLD_DIR / f"{stem}.json"
    if not gold_path.exists():
        return None
    gold = json.loads(gold_path.read_text())
    try:
        return lenient_condition_coverage(pred, gold)
    except Exception as e:
        print(f"  ! lenient failed on {system}/{stem}: {e}", file=sys.stderr)
        return None


def main():
    wb = load_workbook(XLSX_PATH)
    ws_per = wb["GT3_Maarten_per_image"]
    ws_agg = wb["GT3_Maarten_aggregate"]

    # ───── per-image sheet: rescore + tag ─────
    headers = [c.value for c in ws_per[1]]
    col_image  = headers.index("image") + 1
    col_system = headers.index("system") + 1
    col_cprec  = headers.index("cond_precision") + 1
    col_crec   = headers.index("cond_recall") + 1

    # Add cond_metric_mode column if not present.
    if "cond_metric_mode" not in headers:
        new_col = ws_per.max_column + 1
        ws_per.cell(row=1, column=new_col, value="cond_metric_mode")
        col_mode = new_col
    else:
        col_mode = headers.index("cond_metric_mode") + 1

    rescored = updated_cells = 0
    skipped: list[tuple[str, str]] = []
    for row in range(2, ws_per.max_row + 1):
        stem   = ws_per.cell(row=row, column=col_image).value
        system = ws_per.cell(row=row, column=col_system).value
        if not stem or not system:
            continue

        if system in SUITE_FOR_SYSTEM:
            cov = _lenient_for(system, stem)
            if cov is None:
                skipped.append((system, stem))
                ws_per.cell(row=row, column=col_mode, value="lenient_unavailable")
                continue
            ws_per.cell(row=row, column=col_cprec, value=cov["precision"])
            ws_per.cell(row=row, column=col_crec,  value=cov["recall"])
            ws_per.cell(row=row, column=col_mode,  value="lenient")
            rescored += 1
            updated_cells += 2
        elif _is_already_lenient(system):
            ws_per.cell(row=row, column=col_mode, value="lenient")
        else:
            ws_per.cell(row=row, column=col_mode, value="strict")

    # ───── aggregate sheet: recompute means + tag ─────
    agg_headers = [c.value for c in ws_agg[1]]
    col_a_system = agg_headers.index("system") + 1
    col_a_cprec  = agg_headers.index("mean_cond_precision") + 1
    col_a_crec   = agg_headers.index("mean_cond_recall") + 1

    if "cond_metric_mode" not in agg_headers:
        new_col = ws_agg.max_column + 1
        ws_agg.cell(row=1, column=new_col, value="cond_metric_mode")
        col_a_mode = new_col
    else:
        col_a_mode = agg_headers.index("cond_metric_mode") + 1

    # Recompute per-system aggregates from the now-lenient per-image values.
    for row in range(2, ws_agg.max_row + 1):
        system = ws_agg.cell(row=row, column=col_a_system).value
        if not system:
            continue

        # Walk per-image rows for this system, collect lenient cond_*
        precs: list[float] = []
        recs: list[float] = []
        for r in range(2, ws_per.max_row + 1):
            if ws_per.cell(row=r, column=col_system).value != system:
                continue
            mode = ws_per.cell(row=r, column=col_mode).value
            if mode != "lenient":
                continue
            p = ws_per.cell(row=r, column=col_cprec).value
            q = ws_per.cell(row=r, column=col_crec).value
            if isinstance(p, (int, float)):
                precs.append(p)
            if isinstance(q, (int, float)):
                recs.append(q)
        if precs:
            ws_agg.cell(row=row, column=col_a_cprec, value=mean(precs))
        if recs:
            ws_agg.cell(row=row, column=col_a_crec, value=mean(recs))
        ws_agg.cell(row=row, column=col_a_mode,
                    value="lenient" if (precs or recs) else "lenient_unavailable")

    wb.save(XLSX_PATH)

    # ───── report ─────
    print(f"Per-image sheet: rescored {rescored} rows ({updated_cells} cells), tagged mode")
    if skipped:
        print(f"  could not rescore (no prediction available):")
        for sys_, stem in skipped:
            print(f"    {sys_}: {stem}")
    print(f"Aggregate sheet: recomputed means for the 3 retroactively-rescored systems, tagged mode")
    print(f"\nFinal aggregate (all lenient now):")
    for row in range(2, ws_agg.max_row + 1):
        system = ws_agg.cell(row=row, column=col_a_system).value
        cp = ws_agg.cell(row=row, column=col_a_cprec).value
        cr = ws_agg.cell(row=row, column=col_a_crec).value
        mode = ws_agg.cell(row=row, column=col_a_mode).value
        cp_s = f"{cp:.3f}" if isinstance(cp, (int, float)) else str(cp)
        cr_s = f"{cr:.3f}" if isinstance(cr, (int, float)) else str(cr)
        print(f"  {system:<32}  cond_prec={cp_s:>6}  cond_rec={cr_s:>6}  ({mode})")


if __name__ == "__main__":
    main()
