"""Regenerate progress-April figure 1 (Soft / Partial / Constituent F1 bars).

Reads the `Fig2_F1_metrics` sheet of progress_April_data.xlsx and writes
fig1_softmatch_f1.png next to it. Sorted by Partial F1 descending so the
strongest systems sit on the left.
"""
from __future__ import annotations

from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook

FIG_DIR = Path(
    "/Users/kasperhuysentruyt/Library/CloudStorage/OneDrive-UGent/"
    "0.Thesis/current/thesis/progress_April_figures"
)
XLSX = FIG_DIR / "progress_April_data.xlsx"
OUT_PNG = FIG_DIR / "fig1_softmatch_f1.png"

SOFT_BLUE    = "#1E64C8"
PARTIAL_BLUE = "#71A8E8"
CONST_GREY   = "#9A9A9A"
INK          = "#1A1814"


def load_rows():
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["Fig2_F1_metrics"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        rows.append({
            "system":     r[idx["System"]],
            "soft":       r[idx["Soft F1 (strict)"]],
            "partial":    r[idx["Partial F1 (Jaccard >= 0.5)"]],
            "constituent": r[idx["Constituent F1"]],
        })
    rows.sort(key=lambda r: r["partial"], reverse=True)
    return rows


def _wrap_label(s: str) -> str:
    """Two-line system labels to keep the x-axis compact."""
    if s.startswith("SDK "):
        return "SDK\n" + s[len("SDK "):]
    if s == "ChemEagle Hybrid":
        return "ChemEagle\nHybrid"
    return s


def main():
    rows = load_rows()
    systems = [r["system"] for r in rows]
    soft = np.array([r["soft"] for r in rows])
    partial = np.array([r["partial"] for r in rows])
    const = np.array([r["constituent"] for r in rows])

    n = len(systems)
    x = np.arange(n)
    w = 0.27

    fig, ax = plt.subplots(figsize=(13, 5.5), facecolor="white")
    ax.set_facecolor("white")

    ax.bar(x - w, soft,    width=w, color=SOFT_BLUE,    label="Soft F1 (strict)")
    ax.bar(x,     partial, width=w, color=PARTIAL_BLUE, label="Partial F1 (Jaccard ≥ 0.5)")
    ax.bar(x + w, const,   width=w, color=CONST_GREY,   label="Constituent F1")

    ax.set_xticks(x)
    ax.set_xticklabels([_wrap_label(s) for s in systems], fontsize=10, color=INK)
    ax.set_ylabel("F1", fontsize=11, color=INK)
    ax.set_ylim(0, 1.0)
    ax.set_yticks(np.arange(0, 1.01, 0.2))
    ax.set_title(
        "Soft-match metrics on the 16-image GT3_Maarten benchmark",
        fontsize=13, fontweight="bold", color=INK, pad=12,
    )

    ax.grid(True, axis="y", color="#888", alpha=0.25, linewidth=0.5, linestyle="--")
    for spine in ("top", "right"):
        ax.spines[spine].set_visible(False)

    ax.legend(loc="upper right", frameon=True, fontsize=10, framealpha=1.0,
              edgecolor="#CCC")

    fig.tight_layout()
    fig.savefig(OUT_PNG, dpi=180, facecolor="white", bbox_inches="tight")
    print(f"Saved: {OUT_PNG}")
    print("\nOrdering (by Partial F1 desc):")
    for r in rows:
        print(f"  {r['system']:<22}  soft={r['soft']:.2f}  partial={r['partial']:.2f}  const={r['constituent']:.2f}")


if __name__ == "__main__":
    main()
