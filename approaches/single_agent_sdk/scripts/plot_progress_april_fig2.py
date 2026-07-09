"""Regenerate progress-April figure 2 (cost vs Partial F1).

Reads the `Fig3_cost_quality` sheet of progress_April_data.xlsx and writes
fig2_cost_quality.png next to it.
"""
from __future__ import annotations

from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook

FIG_DIR = Path(
    "/Users/kasperhuysentruyt/Library/CloudStorage/OneDrive-UGent/"
    "0.Thesis/current/thesis/progress_April_figures"
)
XLSX = FIG_DIR / "progress_April_data.xlsx"
OUT_PNG = FIG_DIR / "fig2_cost_quality.png"

SDK_BLUE  = "#1E64C8"
BASE_GREY = "#9A9A9A"
INK       = "#1A1814"

LABEL_OFFSETS = {
    "SDK Opus 4.7":         (8, 6),
    "SDK Opus 4.6":         (8, -14),
    "SDK Gemini 3.1 Pro":   (8, -4),
    "SDK Gemini 3 Flash":   (8, 6),
    "SDK Grok 4.3":         (8, 6),
    "SDK GPT-5.4":          (8, 6),
    "SDK GPT-5.5":          (8, 6),
    "ChemEagle":            (-8, 14),
    "ChemEagle Hybrid":     (-8, -16),
    "multi-agent":          (8, -14),
}

# Systems excluded from the log-linear trend fit (still plotted, just not fitted).
# Outliers that flip the trend without representing the cost-quality story:
#   - "SDK Gemini 3 Flash": exceptionally high F1 at exceptionally low cost
#   - "multi-agent": high cost with low F1 (multi-step orchestration overhead)
TREND_EXCLUDE = {"SDK Gemini 3 Flash", "multi-agent"}


def load_rows():
    wb = load_workbook(XLSX, data_only=False)
    ws = wb["Fig3_cost_quality"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        n = r[idx["Images (n/16)"]]
        usd_total = r[idx["USD cost (total)"]]
        rows.append({
            "system":  r[idx["System"]],
            "n":       n,
            "f1":      r[idx["Partial F1"]],
            "usd_per": usd_total / n,
            "category": r[idx["Category"]],
        })
    return rows


def main():
    rows = load_rows()

    fig, ax = plt.subplots(figsize=(11, 5.5), facecolor="white")
    ax.set_facecolor("white")

    drew = {"Single-SDK agent": False, "Multi-step baseline": False}
    for r in rows:
        is_sdk = r["category"] == "Single-SDK agent"
        color = SDK_BLUE if is_sdk else BASE_GREY
        cat = r["category"]
        label = cat if not drew[cat] else None
        drew[cat] = True
        ax.scatter(
            r["usd_per"], r["f1"],
            s=260, c=color, edgecolors="white", linewidths=1.5,
            alpha=0.95, zorder=3, label=label,
        )
        dx, dy = LABEL_OFFSETS.get(r["system"], (8, 6))
        ax.annotate(
            r["system"],
            (r["usd_per"], r["f1"]),
            xytext=(dx, dy), textcoords="offset points",
            fontsize=10, color=INK,
            ha="left" if dx >= 0 else "right",
            va="center",
        )

    # Log-linear trend: F1 = a*log10(usd_per) + b, fit on points NOT in TREND_EXCLUDE.
    fit_rows = [r for r in rows if r["system"] not in TREND_EXCLUDE]
    xs = np.array([r["usd_per"] for r in fit_rows])
    ys = np.array([r["f1"]      for r in fit_rows])
    log_xs = np.log10(xs)
    a, b = np.polyfit(log_xs, ys, 1)
    y_pred = a * log_xs + b
    ss_res = np.sum((ys - y_pred) ** 2)
    ss_tot = np.sum((ys - ys.mean()) ** 2)
    r2 = 1 - ss_res / ss_tot if ss_tot > 0 else float("nan")
    # Draw the trend line over the full plotted x-range so excluded points are visually outside the fit.
    all_xs = np.array([r["usd_per"] for r in rows])
    line_x = np.geomspace(all_xs.min() * 0.8, all_xs.max() * 1.2, 100)
    line_y = a * np.log10(line_x) + b
    excl_short = ", ".join(sorted({s.replace("SDK ", "") for s in TREND_EXCLUDE}))
    ax.plot(line_x, line_y, color=INK, alpha=0.35, linewidth=1.4,
            linestyle="--", zorder=1,
            label=f"Log-linear fit, n={len(fit_rows)} (R²={r2:.2f}; excl. {excl_short})")

    ax.set_xscale("log")
    ax.set_xlabel("Cost per image (USD, log scale)", fontsize=11, color=INK)
    ax.set_ylabel("Partial F1 (Jaccard ≥ 0.5)", fontsize=11, color=INK)
    ax.set_title(
        "Cost vs quality on the 16-image benchmark",
        fontsize=13, fontweight="bold", color=INK, pad=12,
    )

    ax.set_ylim(0.1, 1.0)
    ax.set_yticks([0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0])

    ax.grid(True, which="both", color="#888", alpha=0.25, linewidth=0.5, linestyle="--")
    for spine in ("top", "right"):
        ax.spines[spine].set_visible(False)

    ax.legend(loc="lower right", frameon=True, fontsize=10, framealpha=1.0,
              edgecolor="#CCC")

    fig.tight_layout()
    fig.savefig(OUT_PNG, dpi=180, facecolor="white", bbox_inches="tight")
    print(f"Saved: {OUT_PNG}")
    print("\nPoints:")
    for r in sorted(rows, key=lambda r: r["usd_per"]):
        print(f"  {r['system']:<22}  ${r['usd_per']:.3f}/img   F1={r['f1']:.3f}   ({r['category']})")


if __name__ == "__main__":
    main()
