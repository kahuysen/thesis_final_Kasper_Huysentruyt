"""Plot the GT3_Maarten_aggregate sheet as small-multiples ranked bar charts.

Output: PNG written next to the xlsx and to ./benchmark_runs/.
"""
from __future__ import annotations

import sys
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
COLL = Path("/Users/kasperhuysentruyt/Documents/thesis/5.Random_tests/Collective_autogen")
XLSX_PATH = COLL / "eval" / "results" / "eval_summary.xlsx"

# Metrics to chart, in order. Skipping intrinsic metrics that are ~1.0 for
# every system (smiles_valid_rate, role_enum_rate) and hard_f1 (near-zero
# for everyone — uninformative).
METRICS = [
    ("mean_product_iou",       "Product IoU",          "higher is better"),
    ("mean_reactant_iou",      "Reactant IoU",         "higher is better"),
    ("mean_soft_f1",           "Soft-match F1",        "higher is better"),
    ("mean_partial_f1",        "Partial-match F1 (≥0.5)", "higher is better"),
    ("mean_constitution_f1",   "Constitution F1",      "higher is better"),
    ("mean_cond_recall",       "Condition recall",     "lenient · higher is better"),
    ("mean_cond_precision",    "Condition precision",  "lenient · higher is better"),
    ("n_schema_pass",          "Schema-valid images",  "higher is better"),
]

# Visual highlighting — our systems get the rust accent; others muted greys.
RUST       = "#B5562C"
RUST_DARK  = "#933F1A"
INK        = "#1A1814"
INK_2      = "#5B564D"
PAPER      = "#F7F4ED"
PAPER_2    = "#EFEBDF"
HAIRLINE   = "#1A1814"  # used at low alpha


def load_aggregate() -> tuple[list[str], dict[str, dict[str, float]]]:
    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb["GT3_Maarten_aggregate"]
    headers = [c.value for c in ws[1]]
    out: dict[str, dict[str, float]] = {}
    systems: list[str] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        sys_name = row[0]
        systems.append(sys_name)
        out[sys_name] = {h: v for h, v in zip(headers, row)}
    return systems, out


def fmt_val(v, max_v):
    """Format a number as either percent or two-decimal float."""
    if isinstance(v, (int, float)):
        if max_v <= 1.05:
            return f"{v:.0%}"
        if isinstance(v, int) or v == int(v):
            return f"{int(v)}"
        return f"{v:.2f}"
    return "—"


def main():
    systems, agg = load_aggregate()
    print(f"Loaded {len(systems)} systems from aggregate sheet")
    for s in systems:
        print(f"  - {s}")

    # Layout: 4 rows × 2 cols = 8 panels (we use all 8 metrics).
    fig, axes = plt.subplots(4, 2, figsize=(14, 13.5), facecolor=PAPER)
    fig.suptitle(
        "Single-SDK agent vs. baselines on GT3_Maarten benchmark",
        fontsize=16, fontweight="bold", color=INK, y=0.995,
    )
    fig.text(
        0.5, 0.972,
        f"{len(systems)} systems · 16 figures · means across the benchmark",
        ha="center", va="top", fontsize=10, color=INK_2,
    )

    for ax, (key, title, sub) in zip(axes.flat, METRICS):
        # Pull values, replace None with 0 for plotting but flag in labels
        rows = []
        for s in systems:
            v = agg[s].get(key)
            rows.append((s, v))
        # Sort descending by value (None last)
        rows.sort(key=lambda r: (r[1] if isinstance(r[1], (int, float)) else -1), reverse=True)

        names = [r[0] for r in rows]
        vals = [r[1] if isinstance(r[1], (int, float)) else 0 for r in rows]
        is_ours = [n.startswith("single_sdk_agent") for n in names]

        max_v = max(vals) if vals else 1.0
        bar_color = [RUST if mine else "#B0AAA0" for mine in is_ours]
        edge_color = [RUST_DARK if mine else INK_2 for mine in is_ours]

        y_pos = list(range(len(names)))
        bars = ax.barh(y_pos, vals, color=bar_color, edgecolor=edge_color, linewidth=1)
        ax.set_yticks(y_pos)
        ax.set_yticklabels(names, fontsize=9, color=INK)
        ax.invert_yaxis()
        ax.set_facecolor(PAPER)

        # X axis range: percent metrics 0–1, count metric 0–max
        if max_v <= 1.05:
            ax.set_xlim(0, 1.0)
            ax.set_xticks([0, 0.25, 0.50, 0.75, 1.0])
            ax.set_xticklabels(["0", "25%", "50%", "75%", "100%"], fontsize=8, color=INK_2)
        else:
            ax.set_xlim(0, max(max_v * 1.1, 1))

        # Title + subtitle
        ax.set_title(title, fontsize=11, fontweight="bold", color=INK, loc="left", pad=2)
        ax.text(0.0, 1.06, sub, transform=ax.transAxes, fontsize=8.5,
                color=INK_2, ha="left", va="bottom")

        # Hide top + right spines, fade bottom + left
        for spine in ("top", "right"):
            ax.spines[spine].set_visible(False)
        for spine in ("bottom", "left"):
            ax.spines[spine].set_color(INK_2)
            ax.spines[spine].set_alpha(0.3)
        ax.tick_params(colors=INK_2, length=3)

        # Value labels at bar tips
        for bar, val, mine in zip(bars, vals, is_ours):
            label = fmt_val(val, max_v)
            color = RUST_DARK if mine else INK_2
            weight = "bold" if mine else "normal"
            xpos = bar.get_width()
            offset = (max_v if max_v > 1.05 else 1.0) * 0.012
            ax.text(xpos + offset, bar.get_y() + bar.get_height() / 2,
                    label, va="center", ha="left",
                    fontsize=9, color=color, fontweight=weight)

    # Legend explaining the rust highlight
    legend_handles = [
        plt.Rectangle((0, 0), 1, 1, fc=RUST, ec=RUST_DARK, label="single_sdk_agent (this work)"),
        plt.Rectangle((0, 0), 1, 1, fc="#B0AAA0", ec=INK_2, label="baselines"),
    ]
    fig.legend(handles=legend_handles, loc="lower center",
               ncol=2, fontsize=10, frameon=False, bbox_to_anchor=(0.5, -0.005))

    fig.tight_layout(rect=[0, 0.02, 1, 0.95])

    # Save in two places
    out1 = COLL / "eval" / "results" / "eval_summary_chart.png"
    out2 = ROOT / "benchmark_runs" / "eval_summary_chart.png"
    out2.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out1, dpi=150, facecolor=PAPER, bbox_inches="tight")
    fig.savefig(out2, dpi=150, facecolor=PAPER, bbox_inches="tight")
    print(f"\nSaved:\n  {out1}\n  {out2}")


if __name__ == "__main__":
    main()
