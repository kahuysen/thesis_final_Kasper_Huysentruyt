"""Cost-quality scatter for the single_sdk_agent variants.

X axis: mean wall time per image (seconds, log scale).
Y axis: mean Product IoU.
Bubble size: mean output tokens per image.
Colour: rust accent for the project's runs; UGent yellow border on Pareto-front points.
"""
from __future__ import annotations

import json
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
COLL = ROOT.parent / "collective_autogen"
XLSX_PATH = COLL / "eval" / "results" / "eval_summary.xlsx"
COSTS_PATH = ROOT / "benchmark_runs" / "_costs_summary.json"

# Map run-dir name → display label as in the workbook
DIR_TO_SYSTEM = {
    "run01_opus4.7":            "single_sdk_agent (Opus 4.7)",
    "run_opus46":               "single_sdk_agent (Opus 4.6)",
    "run_gpt54":                "single_sdk_agent (GPT-5.4)",
    "run_gpt55":                "single_sdk_agent (GPT-5.5)",
    "run_gemini31propreview":   "single_sdk_agent (Gemini 3.1 Pro Preview)",
    "run_gemini3flash":         "single_sdk_agent (Gemini 3 Flash Preview)",
    "run_grok43":               "single_sdk_agent (Grok 4.3)",
}

UGENT_BLUE       = "#1E64C8"
UGENT_BLUE_DARK  = "#0A3B7A"
UGENT_YELLOW     = "#FFD200"
RUST             = "#B5562C"
INK              = "#1A1814"
INK_2            = "#5B564D"
PAPER            = "#F7F4ED"


def quality_for(system: str) -> float | None:
    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb["GT3_Maarten_aggregate"]
    hdr = [c.value for c in ws[1]]
    ix = hdr.index("mean_product_iou")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] == system:
            return row[ix]
    return None


def main():
    costs = json.loads(COSTS_PATH.read_text())

    rows = []
    for run_dir, sys_name in DIR_TO_SYSTEM.items():
        c = costs.get(run_dir)
        q = quality_for(sys_name)
        if not c or q is None:
            continue
        rows.append({
            "system":  sys_name,
            "label":   sys_name.replace("single_sdk_agent (", "").rstrip(")"),
            "wall_s":  c["elapsed_s_mean"],
            "in_tok":  c["input_tokens_mean"],
            "out_tok": c["output_tokens_mean"],
            "iou":     q,
        })

    fig, ax = plt.subplots(figsize=(10, 6.5), facecolor="white")
    ax.set_facecolor(PAPER)

    # Draw a faint Pareto-front guide: any point that's both faster and higher-quality
    # than all others is on the frontier.
    sorted_by_speed = sorted(rows, key=lambda r: r["wall_s"])
    pareto = []
    best_q = -1
    for r in sorted_by_speed:
        if r["iou"] > best_q:
            pareto.append(r)
            best_q = r["iou"]

    # Connect Pareto-front points with a thin line
    if len(pareto) >= 2:
        ax.plot(
            [r["wall_s"] for r in pareto],
            [r["iou"]   for r in pareto],
            color=UGENT_YELLOW, linewidth=1.5, alpha=0.6, zorder=1,
            linestyle="--", label="Pareto frontier",
        )

    pareto_set = {r["system"] for r in pareto}
    for r in rows:
        size = max(80, min(1200, r["out_tok"] / 5))   # bubble area ~ output tokens
        on_pareto = r["system"] in pareto_set
        face = RUST
        edge = UGENT_BLUE_DARK if on_pareto else "#888"
        lw = 2.5 if on_pareto else 0.8
        ax.scatter(r["wall_s"], r["iou"], s=size, c=face,
                   edgecolors=edge, linewidths=lw, alpha=0.85, zorder=3)
        # Label
        offset_x = r["wall_s"] * 0.06
        ax.annotate(
            r["label"],
            (r["wall_s"], r["iou"]),
            xytext=(offset_x + 4, 0), textcoords="offset points",
            fontsize=10, color=INK, va="center",
            fontweight="bold" if on_pareto else "normal",
        )

    ax.set_xscale("log")
    ax.set_xlabel("Mean wall time per image (s, log scale)", fontsize=11, color=INK)
    ax.set_ylabel("Mean Product IoU", fontsize=11, color=INK)

    fig.suptitle(
        "Cost-quality Pareto on the GT3_Maarten benchmark",
        fontsize=14, fontweight="bold", color=INK, y=0.98,
    )
    fig.text(
        0.5, 0.93,
        "Bubble area ∝ mean output tokens / image  ·  yellow dashed = Pareto frontier",
        ha="center", va="top", fontsize=9, color=INK_2,
    )

    # Light grid
    ax.grid(True, which="both", color="#1A1814", alpha=0.08, linewidth=0.5)
    for spine in ("top", "right"):
        ax.spines[spine].set_visible(False)
    for spine in ("bottom", "left"):
        ax.spines[spine].set_color(INK_2)
        ax.spines[spine].set_alpha(0.3)
    ax.tick_params(colors=INK_2)

    ax.set_ylim(0.0, 1.0)
    # Pad the x-axis on the right so the "Grok 4.3" label has room
    ax.set_xlim(ax.get_xlim()[0], ax.get_xlim()[1] * 1.6)
    ax.legend(loc="lower right", frameon=False, fontsize=9)

    out1 = COLL / "eval" / "results" / "cost_quality.png"
    out2 = ROOT / "benchmark_runs" / "cost_quality.png"
    out2.parent.mkdir(parents=True, exist_ok=True)
    fig.tight_layout(rect=[0, 0, 1, 0.91])
    fig.savefig(out1, dpi=180, facecolor="white", bbox_inches="tight")
    fig.savefig(out2, dpi=180, facecolor="white", bbox_inches="tight")
    print(f"Saved:\n  {out1}\n  {out2}")
    print(f"\nPareto frontier: {[r['label'] for r in pareto]}")


if __name__ == "__main__":
    main()
