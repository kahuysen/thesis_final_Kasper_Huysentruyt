"""Heatmap of GT3_Maarten_aggregate in UGent brand colors.

Color scale: white → UGent blue (#1E64C8 → #0A3B7A) per-column normalised
so the visual ranks each metric independently. Numeric values printed on
each cell are absolute. Our two systems' row labels are marked with the
UGent yellow accent (#FFD200) so they pop without polluting the data layer.

Output: PNG next to the xlsx and in the project's benchmark_runs/.
"""
from __future__ import annotations

import sys
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.colors import LinearSegmentedColormap
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
COLL = ROOT.parent / "collective_autogen"
XLSX_PATH = COLL / "eval" / "results" / "eval_summary.xlsx"

# UGent brand palette
UGENT_BLUE       = "#1E64C8"
UGENT_BLUE_DARK  = "#0A3B7A"
UGENT_BLUE_PALE  = "#F4F7FD"
UGENT_YELLOW     = "#FFD200"
UGENT_YELLOW_DARK = "#C7A700"
INK              = "#1A1814"
INK_2            = "#5B564D"

# Metrics to show (column key, display label)
METRICS = [
    ("mean_product_iou",      "Product IoU"),
    ("mean_reactant_iou",     "Reactant IoU"),
    ("mean_soft_f1",          "Soft F1"),
    ("mean_partial_f1",       "Partial F1 (≥0.5)"),
    ("mean_constitution_f1",  "Constitution F1"),
    ("mean_cond_recall",      "Cond. recall (lenient)"),
    ("mean_cond_precision",   "Cond. precision (lenient)"),
    ("n_schema_pass",         "Schema-valid (of 16)"),
]


def load_aggregate():
    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb["GT3_Maarten_aggregate"]
    headers = [c.value for c in ws[1]]
    systems, rows = [], {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        systems.append(row[0])
        rows[row[0]] = {h: v for h, v in zip(headers, row)}
    return systems, rows


def build_matrix(systems: list[str], rows: dict, metrics: list[tuple[str, str]]):
    """Returns (raw_values: 2D, normalised: 2D-in-[0,1], display_strings: 2D)."""
    raw = np.full((len(systems), len(metrics)), np.nan)
    for i, s in enumerate(systems):
        for j, (k, _) in enumerate(metrics):
            v = rows[s].get(k)
            if isinstance(v, (int, float)):
                raw[i, j] = v
    # Per-column min-max normalisation, NaN-safe
    norm = np.zeros_like(raw)
    for j in range(raw.shape[1]):
        col = raw[:, j]
        if np.all(np.isnan(col)):
            continue
        cmin = np.nanmin(col)
        cmax = np.nanmax(col)
        rng = cmax - cmin
        if rng == 0:
            norm[:, j] = 1.0  # all equal → uniform top-shade (avoid div-zero)
        else:
            norm[:, j] = (col - cmin) / rng
    # Pretty cell text
    disp = np.empty(raw.shape, dtype=object)
    for i in range(raw.shape[0]):
        for j in range(raw.shape[1]):
            v = raw[i, j]
            if np.isnan(v):
                disp[i, j] = "—"
            elif metrics[j][0] == "n_schema_pass":
                disp[i, j] = f"{int(v)}"
            elif abs(v) <= 1.0:
                disp[i, j] = f"{v:.0%}"
            else:
                disp[i, j] = f"{v:.2f}"
    return raw, norm, disp


def main():
    systems, rows = load_aggregate()
    raw, norm, disp = build_matrix(systems, rows, METRICS)

    cmap = LinearSegmentedColormap.from_list(
        "ugent_seq",
        [UGENT_BLUE_PALE, UGENT_BLUE, UGENT_BLUE_DARK],
        N=256,
    )

    n_rows, n_cols = norm.shape
    fig, ax = plt.subplots(figsize=(13, 5.6), facecolor="white")

    im = ax.imshow(norm, cmap=cmap, aspect="auto", vmin=0, vmax=1)

    # Move column labels to the top + rotate 30° so longer labels never collide.
    ax.xaxis.set_ticks_position("top")
    ax.xaxis.set_label_position("top")
    ax.set_xticks(range(n_cols))
    ax.set_xticklabels([m[1] for m in METRICS], fontsize=10, color=INK,
                       rotation=30, ha="left", rotation_mode="anchor")
    ax.set_yticks(range(n_rows))
    ax.set_yticklabels(systems, fontsize=11, color=INK)

    # Highlight our two rows with the UGent yellow accent (label color + tick mark)
    for tick_label, sys_name in zip(ax.get_yticklabels(), systems):
        if sys_name.startswith("single_sdk_agent"):
            tick_label.set_color(UGENT_YELLOW_DARK)
            tick_label.set_fontweight("bold")

    # Cell text — auto-pick black or white per cell brightness for legibility
    for i in range(n_rows):
        for j in range(n_cols):
            shade = norm[i, j]
            text_color = "white" if shade > 0.55 else INK
            ax.text(j, i, disp[i, j], ha="center", va="center",
                    fontsize=10, color=text_color,
                    fontweight="bold" if systems[i].startswith("single_sdk_agent") else "normal")

    # Yellow highlight stripe on the left of our rows
    for i, s in enumerate(systems):
        if s.startswith("single_sdk_agent"):
            ax.add_patch(plt.Rectangle((-0.55, i - 0.5), 0.1, 1.0,
                                        color=UGENT_YELLOW, clip_on=False, zorder=10))

    # Hide spines + ticks
    for spine in ax.spines.values():
        spine.set_visible(False)
    ax.tick_params(length=0)
    ax.set_xticks(np.arange(-0.5, n_cols, 1), minor=True)
    ax.set_yticks(np.arange(-0.5, n_rows, 1), minor=True)
    ax.grid(which="minor", color="white", linewidth=2)
    ax.tick_params(which="minor", length=0)

    # Caption sits at the bottom now that column labels live up top.
    fig.text(
        0.5, 0.02,
        "Color: per-column min-max ranking (light = worst in column, dark = best).   "
        "Cell value: absolute metric.   "
        "Yellow stripe + bold gold label = single_sdk_agent (this work).",
        ha="center", va="bottom", fontsize=9, color=INK_2,
    )

    # Colorbar — explains the color = "rank within column"
    cbar = fig.colorbar(im, ax=ax, fraction=0.025, pad=0.02)
    cbar.set_label("per-column rank", fontsize=8.5, color=INK_2)
    cbar.ax.tick_params(labelsize=7, colors=INK_2, length=2)
    cbar.set_ticks([0, 0.5, 1.0])
    cbar.set_ticklabels(["worst", "mid", "best"])
    cbar.outline.set_visible(False)

    fig.suptitle(
        "GT3_Maarten benchmark — system × metric heatmap",
        fontsize=15, fontweight="bold", color=INK, y=1.04,
    )
    fig.tight_layout(rect=[0, 0.05, 1, 0.99])

    out1 = COLL / "eval" / "results" / "eval_summary_heatmap_ugent.png"
    out2 = ROOT / "benchmark_runs" / "eval_summary_heatmap_ugent.png"
    out2.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out1, dpi=180, bbox_inches="tight", facecolor="white")
    fig.savefig(out2, dpi=180, bbox_inches="tight", facecolor="white")
    print(f"Saved:\n  {out1}\n  {out2}")


if __name__ == "__main__":
    main()
