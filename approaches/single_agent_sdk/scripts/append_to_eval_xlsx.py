"""Append our two runs (Opus 4.7, GPT-5.4) to the Collective_autogen
`eval_summary.xlsx`, using the same metric columns and system-name convention
as the existing rows (multi_agent, ChemEagle, ChemEagle_Hybrid, …).

System name format: 'single_sdk_agent (<model>)'.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path
from statistics import mean

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

COLL = ROOT.parent / "collective_autogen"
sys.path.insert(0, str(COLL))
sys.path.insert(0, str(ROOT / "scripts"))

from openpyxl import load_workbook
from eval.metrics import evaluate
from eval_via_collective import figure_extraction_to_record, lenient_condition_coverage

GOLD_DIR = COLL / "eval" / "ground_truth"
XLSX_PATH = COLL / "eval" / "results" / "eval_summary.xlsx"

PER_IMAGE_COLS = [
    "image", "system", "schema_pass",
    "n_reactions_pred", "n_reactions_gold",
    "smiles_valid_rate", "role_enum_rate",
    "reactant_iou", "product_iou",
    "cond_precision", "cond_recall",
    "soft_f1", "hard_f1", "constitution_f1", "partial_f1",
    "ged",
]

AGG_COLS = [
    "system", "n_images_scored", "n_schema_pass",
    "mean_smiles_valid_rate", "mean_role_enum_rate",
    "mean_reactant_iou", "mean_product_iou",
    "mean_cond_precision", "mean_cond_recall",
    "mean_soft_f1", "mean_hard_f1",
    "mean_constitution_f1", "mean_partial_f1", "mean_ged",
]


def per_image_row(image_stem: str, system: str, m: dict) -> tuple:
    """Translate one evaluate() result dict into the existing sheet's column order.

    cond_precision / cond_recall use the LENIENT matcher (paren stripping,
    number-unit whitespace collapse, role bucketing, SMILES-based unification).
    See pipeline/eval_via_collective.py:lenient_condition_coverage.
    """
    def _get(parent, child, default=None):
        v = m.get(parent, {}).get(child) if parent in m else default
        return v if v is not None else default

    cond = m.get("condition_coverage_lenient") or m.get("condition_coverage", {})

    return (
        image_stem,
        system,
        bool(m["schema_conformance"]["valid"]),
        m["reaction_count"]["count"],
        _get("reaction_count_gold", "count", 0),
        m["smiles_validity_rate"]["rate"],
        m["role_enum_compliance"]["rate"],
        _get("reactant_iou", "iou"),
        _get("product_iou", "iou"),
        cond.get("precision"),
        cond.get("recall"),
        _get("soft_match", "f1"),
        _get("hard_match", "f1"),
        _get("constitution_match", "f1"),
        _get("partial_match", "f1"),
        _get("graph_edit_distance", "ged_per_reaction"),
    )


def aggregate_row(system: str, rows: list[tuple]) -> tuple:
    """Build the aggregate row from already-computed per-image rows."""
    n = len(rows)
    n_pass = sum(1 for r in rows if r[2])

    def _macro(idx):
        vals = [r[idx] for r in rows if isinstance(r[idx], (int, float))]
        return mean(vals) if vals else None

    return (
        system,
        n,
        n_pass,
        _macro(5),   # smiles_valid_rate
        _macro(6),   # role_enum_rate
        _macro(7),   # reactant_iou
        _macro(8),   # product_iou
        _macro(9),   # cond_precision
        _macro(10),  # cond_recall
        _macro(11),  # soft_f1
        _macro(12),  # hard_f1
        _macro(13),  # constitution_f1
        _macro(14),  # partial_f1
        _macro(15),  # ged
    )


def collect_run_rows(run_dir: Path, system: str) -> list[tuple]:
    rows: list[tuple] = []
    for pred_path in sorted(run_dir.glob("*.json")):
        if pred_path.stem.endswith(("_insight", ".original")):
            continue
        gold_path = GOLD_DIR / f"{pred_path.stem}.json"
        if not gold_path.exists():
            print(f"  ! no gold for {pred_path.stem}, skipping", file=sys.stderr)
            continue
        fx = json.loads(pred_path.read_text())
        gold = json.loads(gold_path.read_text())
        record = figure_extraction_to_record(fx, file_name=f"{pred_path.stem}.png")
        try:
            metrics = evaluate(record, gold)
        except Exception as e:
            print(f"  ! evaluate failed on {pred_path.stem}: {e}", file=sys.stderr)
            continue
        try:
            metrics["condition_coverage_lenient"] = lenient_condition_coverage(record, gold)
        except Exception as e:
            print(f"  ! lenient cond coverage failed on {pred_path.stem}: {e}", file=sys.stderr)
        rows.append(per_image_row(pred_path.stem, system, metrics))
    return rows


def main():
    runs = [
        ("benchmark_runs/run01_opus4.7",             "single_sdk_agent (Opus 4.7)"),
        ("benchmark_runs/run_gpt54",                  "single_sdk_agent (GPT-5.4)"),
        ("benchmark_runs/run_gemini31propreview",     "single_sdk_agent (Gemini 3.1 Pro Preview)"),
        ("benchmark_runs/run_gemini3flash",           "single_sdk_agent (Gemini 3 Flash Preview)"),
        ("benchmark_runs/run_grok43",                 "single_sdk_agent (Grok 4.3)"),
    ]

    wb = load_workbook(XLSX_PATH)
    ws_per = wb["GT3_Maarten_per_image"]
    ws_agg = wb["GT3_Maarten_aggregate"]

    # Sanity: confirm the expected columns are at the start of the sheet header.
    # Trailing extras (e.g. `cond_metric_mode` added by relax_cond_in_xlsx.py) are fine.
    header_per = tuple(c.value for c in ws_per[1])
    header_agg = tuple(c.value for c in ws_agg[1])
    assert header_per[:len(PER_IMAGE_COLS)] == tuple(PER_IMAGE_COLS), \
        f"per-image header prefix mismatch:\n  expected {PER_IMAGE_COLS}\n  found    {header_per}"
    assert header_agg[:len(AGG_COLS)] == tuple(AGG_COLS), \
        f"aggregate header prefix mismatch:\n  expected {AGG_COLS}\n  found    {header_agg}"

    # If extra trailing cols exist, pad each appended row with Nones to keep alignment.
    n_trail_per = len(header_per) - len(PER_IMAGE_COLS)
    n_trail_agg = len(header_agg) - len(AGG_COLS)

    # Drop any prior rows we wrote (idempotent re-run)
    def _drop_system(ws, system_col_idx, system):
        # collect rows to delete bottom-up
        to_delete = [
            row for row in range(ws.max_row, 1, -1)
            if ws.cell(row=row, column=system_col_idx).value == system
        ]
        for row in to_delete:
            ws.delete_rows(row)

    for _, system in runs:
        _drop_system(ws_per, system_col_idx=2, system=system)  # 'system' is col 2
        _drop_system(ws_agg, system_col_idx=1, system=system)  # 'system' is col 1

    # Append fresh rows
    for run_path, system in runs:
        run_dir = ROOT / run_path
        print(f"\nScoring {system}  ({run_dir})")
        per_rows = collect_run_rows(run_dir, system)
        for r in per_rows:
            ws_per.append(tuple(r) + (None,) * n_trail_per)
        agg = aggregate_row(system, per_rows)
        ws_agg.append(tuple(agg) + (None,) * n_trail_agg)
        print(f"  → {len(per_rows)} per-image rows + 1 aggregate row")
        print(f"    aggregate: {dict(zip(AGG_COLS, agg))}")

    wb.save(XLSX_PATH)
    print(f"\nSaved to {XLSX_PATH}")


if __name__ == "__main__":
    main()
